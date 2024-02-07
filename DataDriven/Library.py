import json
import jsonpath
import pytest
import openpyxl

class Common:
    def __init__(self,FileNamePath, SheetName):
        global wk
        global sh
        wk = openpyxl.load_workbook(FileNamePath)
        sh = wk[SheetName]


    def abholen_row_count(self):
        rows = sh.max_row
        return rows

    def abholen_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        cols = sh.max_column
        li =[]
        for i in range (1,cols+1):
            cell = sh.cell(row= 1, column=i)
            li.insert(i-1,cell.value)
        return li

    def update_request_mit_data(self,rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1,c+1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]]=cell.value

        return jsonRequest

