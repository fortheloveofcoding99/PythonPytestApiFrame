import json
import jsonpath
import requests
import openpyxl
from DataDriven import Library

def test_add_multi_user():
    app_url = 'https://thetestingworldapi.com/api/studentsDetails'
    file_read = open('../datei/datei1.json')
    json_request = json.loads(file_read.read())


    obj = Library.Common('../datei/Book1.xlsx','Sheet1')
    col = obj.abholen_column_count()
    keyList = obj.fetch_key_names()
    row = obj.abholen_row_count()

    for i in range(2, row + 1):
        updated_json_request = obj.update_request_mit_data(row,json_request,keyList)
        response = requests.post(app_url,updated_json_request)
        print(response.text)
        print(response.status_code)

    # for i in range(2,rows+1):
    #     cell_first_name = sh.cell(row=i,column=1)
    #     cell_mid_name = sh.cell(row=i,column=2)
    #     cell_last_name = sh.cell(row=i,column=3)
    #     cell_dob = sh.cell(row=i,column=4)
    #     json_request['first_name'] = cell_first_name.value
    #     json_request['middle_name'] = cell_mid_name.value
    #     json_request['last_name'] = cell_last_name.value
    #     json_request['date_of_birth'] = cell_dob.value
    #     response1 = requests.get(app_url, json_request)
    #     print(response1.status_code)
    #     print(response1.text)
    #     assert response1.status_code == 200

